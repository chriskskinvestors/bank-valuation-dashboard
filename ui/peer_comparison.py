"""
Peer Comparison UI — side-by-side bank comparison with percentile color-coding.

Top-level section in sidebar nav.
"""

import html as _html

import streamlit as st
import pandas as pd

from config import METRICS_BY_KEY
from data.bank_mapping import get_name
from analysis.peer_groups import (
    asset_size_tier, business_mix_tier, compute_peer_percentile,
)
from utils.formatting import format_value
from ui.chrome import table_export, status_dot


# Curated metric sets per category (focused on what analysts actually look at).
# All keys validated against METRICS_BY_KEY at module load.
CATEGORY_METRICS = {
    "Valuation": [
        "price", "pe_ratio", "ptbv_ratio", "dividend_yield",
        "fair_ptbv", "ptbv_discount", "market_cap",
    ],
    "Profitability": [
        "roaa", "roatce", "roatce_4q", "nim", "nim_4q",
        "efficiency_ratio", "pretax_roa", "net_op_income_assets",
    ],
    "Credit": [
        "npl_ratio", "npl_cre", "nco_ratio",
        "past_due_30_89", "reserve_coverage_pct", "reserve_to_loans",
        "nco_4q_trend_bps", "credit_alerts_count",
    ],
    "Capital": [
        "cet1_ratio", "cet1_current", "cet1_qoq_pp",
        "total_capital_ratio", "leverage_ratio",
        "equity_to_assets", "tbv_cagr_1y",
        "payout_ratio_4q", "buyback_capacity_usd",
    ],
    "Deposits": [
        "total_deposits", "nonint_dep_pct", "uninsured_pct",
        "core_dep_pct", "brokered_pct",
        "deposit_cycle_beta", "deposit_rolling_beta",
        "dep_qoq_growth", "cod_qoq_bps",
    ],
    "Balance Sheet": [
        "total_assets", "total_loans", "loans_to_deposits",
        "ln_cre_pct", "ln_ci_pct", "ln_resi_pct", "ln_consumer_pct",
        "cre_to_capital", "sec_to_assets_pct", "htm_pct",
    ],
}


# Percentile color scale — single source of truth for BOTH the table cells and
# the legend (they previously diverged: the legend showed an old palette that no
# longer appeared anywhere in the table).
# (min_effective_percentile, label, background, text color, bold)
_PCT_SCALE = [
    (80, "Top 20%", "#d1fae5", "#065f46", True),
    (60, "60–80th", "#ecfdf5", "#047857", False),
    (40, "40–60th", "#f8fafc", "#475569", False),
    (20, "20–40th", "#fef2f2", "#991b1b", False),
    (0,  "Bottom 20%", "#fee2e2", "#991b1b", True),
]


def _percentile_color(pct: float | None, higher_better: bool = True) -> str:
    """Color based on percentile rank (0-100)."""
    if pct is None:
        return ""
    # Normalize: if lower-is-better, invert the percentile
    effective = pct if higher_better else (100 - pct)
    for floor, _label, bg, fg, bold in _PCT_SCALE:
        if effective >= floor:
            weight = " font-weight:600;" if bold else ""
            return f"background-color: {bg}; color:{fg};{weight}"
    return ""


# Best-in-class call-outs across the compared banks (label, metric_key, min|max).
_HIGHLIGHTS = [
    ("Cheapest P/TBV", "ptbv_ratio", "min"),
    ("Biggest discount", "ptbv_discount", "min"),
    ("Highest ROATCE", "roatce_normalized", "max"),
    ("Best efficiency", "efficiency_ratio", "min"),
    ("Highest NIM", "nim", "max"),
    ("Strongest CET1", "cet1_ratio", "max"),
    ("Cleanest credit", "npl_ratio", "min"),
    ("Fastest TBV growth", "tbv_cagr_1y", "max"),
]

# Column-ordering choices for the side-by-side table (metric_key, label).
# "__score__" = the composite percentile; the rest order by a single metric.
_RANK_OPTIONS = [("__score__", "Overall score")] + [
    (k, lbl) for k, lbl in [
        ("roatce_normalized", "ROATCE"), ("nim", "NIM"),
        ("efficiency_ratio", "Efficiency"), ("cet1_ratio", "CET1"),
        ("ptbv_ratio", "P/TBV"), ("ptbv_discount", "Discount to fair"),
        ("npl_ratio", "NPL"), ("market_cap", "Market cap"),
        ("total_assets", "Total assets"),
    ] if k in METRICS_BY_KEY
]


def _peer_scores(cohort: list[dict], display_peers: list[dict],
                 categories: list[str]) -> dict:
    """Per-displayed-bank composite: mean EFFECTIVE percentile (higher = better;
    lower-is-better metrics inverted) across every directional metric in the
    selected categories, plus top-/bottom-quartile finish counts. Percentiles
    resolve against the FULL cohort."""
    agg = {d["ticker"]: {"vals": [], "top": 0, "bot": 0} for d in display_peers}
    for cat in categories:
        for mkey in CATEGORY_METRICS.get(cat, []):
            m_def = METRICS_BY_KEY.get(mkey)
            if not m_def:
                continue
            rule = m_def.get("color_rule")
            if rule not in ("higher_better", "lower_better"):
                continue
            numeric = [b.get(mkey) for b in cohort
                       if isinstance(b.get(mkey), (int, float))]
            if not numeric:
                continue
            for d in display_peers:
                v = d.get(mkey)
                if not isinstance(v, (int, float)):
                    continue
                pct = compute_peer_percentile(v, numeric)
                if pct is None:
                    continue
                eff = pct if rule == "higher_better" else (100 - pct)
                a = agg[d["ticker"]]
                a["vals"].append(eff)
                if eff >= 80:
                    a["top"] += 1
                elif eff < 20:
                    a["bot"] += 1
    return {tk: {"score": (sum(a["vals"]) / len(a["vals"]) if a["vals"] else None),
                 "top": a["top"], "bot": a["bot"], "n": len(a["vals"])}
            for tk, a in agg.items()}


def _render_scorecard(scores: dict, order: list[str]):
    """A strip of per-bank chips — composite score + top/bottom-quartile counts —
    in the current column order (best-first when ranked by score)."""
    chips = []
    for tk in order:
        s = scores.get(tk)
        if not s or s.get("score") is None:
            continue
        parts = [f'<b>{_html.escape(tk)}</b>',
                 f'<span style="color:var(--brand-primary);">{round(s["score"])}</span>']
        if s["top"]:
            parts.append(f'<span style="color:#047857;">▲{s["top"]}</span>')
        if s["bot"]:
            parts.append(f'<span style="color:#b91c1c;">▼{s["bot"]}</span>')
        chips.append(
            f'<span style="font-size:0.75rem;background:var(--bg-surface);'
            f'border:0.5px solid var(--grid-head);border-radius:0;padding:4px 9px;'
            f'white-space:nowrap;">{" ".join(parts)}</span>')
    if not chips:
        return
    st.markdown(
        '<div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;'
        'margin:2px 0 10px;">'
        '<span style="font-size:0.72rem;color:var(--text-tertiary);'
        'text-transform:uppercase;letter-spacing:.04em;">Scorecard</span>'
        + "".join(chips)
        + '<span style="font-size:0.68rem;color:var(--text-tertiary);">'
        'score = avg percentile · ▲ top-quartile · ▼ bottom-quartile finishes</span>'
        '</div>',
        unsafe_allow_html=True,
    )


def render_peer_comparison(all_metrics: list[dict]):
    """Render the Peer Comparison page — dense, all-category side-by-side."""
    if not all_metrics:
        st.warning("No bank data loaded. Check your watchlist.")
        return

    from ui.bank_scope import scope_type_options, render_scope_sub

    # Screen → Compare handoff: a "Compare these banks" click on a screen stashes
    # the result set; pre-seed Manual scope with them on arrival, then consume it.
    handoff = st.session_state.pop("_compare_handoff_tickers", None)
    if handoff:
        st.session_state["compare_scope_type"] = "Manual"
        st.session_state["compare_manual"] = list(handoff)

    # Quick "add a bank by ticker" → append to the Manual scope. Handled BEFORE the
    # scope/manual widgets render so writing their session keys is legal; the search
    # box then resets for the next add (mirrors the Screen toolbar's Add bank).
    _addbank = st.session_state.get("compare_addbank")
    if _addbank:
        _cur = list(st.session_state.get("compare_manual", []))
        if _addbank not in _cur:
            _cur.append(_addbank)
        st.session_state["compare_manual"] = _cur
        st.session_state["compare_scope_type"] = "Manual"
        st.session_state["compare_addbank"] = None

    # ── Compact controls: Scope · Categories (content-width via trailing spacer);
    # the scope secondary picker (Manual chips / cohort) sits narrow below. ──
    _CATS = list(CATEGORY_METRICS.keys())
    _all_tk = sorted({m["ticker"] for m in all_metrics if m.get("ticker")})
    cc1, cc2, cc3, cc4, _csp = st.columns([1.3, 2.0, 1.5, 1.9, 1.3])
    with cc1:
        scope_type = st.selectbox("Scope", scope_type_options(), key="compare_scope_type")
    with cc2:
        categories = st.multiselect("Metric categories", _CATS, default=_CATS,
                                    key="peer_categories")
    with cc3:
        rank_label = st.selectbox("Rank banks by", [lbl for _, lbl in _RANK_OPTIONS],
                                  key="compare_rank")
    with cc4:
        st.selectbox(
            "Add bank", options=_all_tk, index=None, placeholder="ticker or name…",
            format_func=lambda t: (f"{t} — {get_name(t)}"
                                   if get_name(t) and get_name(t) != t else t),
            key="compare_addbank")
    if not categories:
        categories = _CATS
    _subl, _ = st.columns([7, 3])
    with _subl:
        cohort, _cohort_tk, peer_label = render_scope_sub(
            all_metrics, scope_type, key_prefix="compare")

    if not cohort:
        st.info("Pick a scope above to compare — a saved group, a tier, a state, "
                "or a manual set (add banks by ticker).")
        return
    if len(cohort) < 2:
        st.warning(f"Only {len(cohort)} bank in this scope — widen it to compare.")

    # Banks are COLUMNS, so the table caps at a readable count; percentiles and the
    # median still resolve against the FULL cohort. Scatter/radar use all of it.
    TABLE_CAP = 12
    cohort_tickers = [m["ticker"] for m in cohort]
    if len(cohort) > TABLE_CAP:
        _dl, _ = st.columns([7, 3])
        with _dl:
            disp_tickers = st.multiselect(
                f"Banks to tabulate (of {len(cohort)}; percentiles vs the full scope)",
                cohort_tickers, default=cohort_tickers[:TABLE_CAP],
                max_selections=TABLE_CAP, key="compare_display")
        display_peers = [m for m in cohort if m["ticker"] in disp_tickers] or cohort[:TABLE_CAP]
    else:
        display_peers = cohort

    # Composite percentile score per bank, then reorder the columns by the rank
    # choice (default: best overall first) so the side-by-side reads at a glance.
    scores = _peer_scores(cohort, display_peers, categories)
    rank_key = next((k for k, lbl in _RANK_OPTIONS if lbl == rank_label), "__score__")
    if rank_key == "__score__":
        display_peers = sorted(
            display_peers,
            key=lambda d: (scores[d["ticker"]]["score"] is None,
                           -(scores[d["ticker"]]["score"] or 0)))
    else:
        _rb_lower = METRICS_BY_KEY.get(rank_key, {}).get("color_rule") == "lower_better"

        def _rank_sort(d):
            v = d.get(rank_key)
            missing = not isinstance(v, (int, float))
            val = v if isinstance(v, (int, float)) else 0
            return (missing, val if _rb_lower else -val)

        display_peers = sorted(display_peers, key=_rank_sort)

    # ── Status line (no boxed title bar) ───────────────────────────────
    _ncat = len(categories)
    st.markdown(
        f'<div style="font-size:var(--fs-xs);color:var(--text-secondary);'
        f'margin:6px 0 8px;">{status_dot("ok", f"{len(cohort)} banks")} · '
        f'{_html.escape(peer_label)} · {len(display_peers)} shown · '
        f'{_ncat} categor{"y" if _ncat == 1 else "ies"} · '
        f'percentiles vs full scope</div>',
        unsafe_allow_html=True,
    )

    # ── Per-bank scorecard + best-in-class highlights ──────────────────
    _render_scorecard(scores, [d["ticker"] for d in display_peers])
    _render_highlights(display_peers)

    # ── View tabs ──────────────────────────────────────────────────────
    view_tab, scatter_tab, radar_tab = st.tabs([
        "Metrics Table", "Scatter Plots", "Radar Chart"])
    with view_tab:
        _cell_mode = st.radio(
            "Cell display", ["Value", "Δ vs median"], horizontal=True,
            key="compare_cellmode", label_visibility="collapsed")
        _render_metrics_table(cohort, display_peers, categories, scores,
                              delta=_cell_mode.startswith("Δ"))
    with scatter_tab:
        _render_peer_scatters(cohort)
    with radar_tab:
        _render_peer_radar(cohort)

    # ── Peer group composition ─────────────────────────────────────────
    st.markdown("---")
    with st.expander("Peer group composition"):
        comp_rows = []
        for m in cohort:
            # total_assets is always raw dollars (converted at the metrics
            # boundary) — no unit guessing.
            assets = m.get("total_assets")
            comp_rows.append({
                "Ticker": m["ticker"],
                "Bank": get_name(m["ticker"]),
                "Assets": format_value(assets, "dollars_auto", 2),
                "Size Tier": asset_size_tier(assets) or "—",
                "Business Mix": business_mix_tier(m),
            })
        comp_df = pd.DataFrame(comp_rows)
        st.dataframe(comp_df, use_container_width=True, hide_index=True)
        table_export(comp_df, "peer_group_composition",
                     key="exp_peer_group_composition")


def _render_highlights(peers: list[dict]):
    """A row of best-in-class chips across the compared banks (no-op for <2)."""
    if len(peers) < 2:
        return
    chips = []
    for label, mkey, mode in _HIGHLIGHTS:
        m_def = METRICS_BY_KEY.get(mkey)
        if not m_def:
            continue
        cand = [(p["ticker"], p.get(mkey)) for p in peers
                if isinstance(p.get(mkey), (int, float))]
        if not cand:
            continue
        tk, val = (min if mode == "min" else max)(cand, key=lambda x: x[1])
        vs = _html.escape(format_value(val, m_def.get("format", "number"),
                                       m_def.get("decimals", 2)))
        chips.append(
            f'<span style="font-size:0.75rem;background:var(--bg-surface);'
            f'border:0.5px solid var(--grid-head);border-radius:0;padding:4px 9px;'
            f'white-space:nowrap;">'
            f'<span style="color:var(--text-secondary);">{_html.escape(label)}</span> '
            f'<b>{_html.escape(tk)}</b> '
            f'<span style="color:var(--brand-primary);">{vs}</span></span>')
    if chips:
        st.markdown(
            f'<div style="display:flex;flex-wrap:wrap;gap:8px;margin:2px 0 12px;">'
            f'{"".join(chips)}</div>',
            unsafe_allow_html=True)


# Headline metrics drawn as bar charts in the right rail (key, label, higher_better).
_HEADLINE_CHARTS = [
    ("roatce_normalized", "ROATCE", True),
    ("nim", "NIM", True),
    ("efficiency_ratio", "Efficiency", False),
    ("cet1_ratio", "CET1", True),
    ("ptbv_ratio", "P/TBV", False),
    ("npl_ratio", "NPL", False),
]


def _sparkline_svg(vals: list, w: int = 46, h: int = 13, color: str = "#1e40af") -> str:
    """Inline-SVG sparkline from a numeric series (None-aware); '' for <2 points."""
    idx = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float))]
    if len(idx) < 2:
        return ""
    nums = [v for _, v in idx]
    lo, hi = min(nums), max(nums)
    rng = (hi - lo) or 1.0
    n = len(vals)
    pts = " ".join(
        f"{(i / (n - 1)) * w:.1f},{h - ((v - lo) / rng) * h:.1f}" for i, v in idx)
    return (f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" '
            f'preserveAspectRatio="none" style="vertical-align:middle;flex:none;">'
            f'<polyline points="{pts}" fill="none" stroke="{color}" '
            f'stroke-width="1.3" stroke-linejoin="round"/></svg>')


def _render_headline_charts(display_peers: list[dict]):
    """Compact horizontal bar charts (one per headline metric) for the compared
    banks — best bank navy, peer median noted. Each FDIC-sourced metric also gets
    an 8-quarter sparkline per bank (from the warm FDIC history cache); computed/
    market metrics show just the bar. Lightweight HTML/SVG (no plotly)."""
    if len(display_peers) < 2:
        return
    from data.loaders import load_fdic_hist
    hist = {}
    for d in display_peers:
        try:
            hist[d["ticker"]] = load_fdic_hist(d["ticker"]) or []
        except Exception:
            hist[d["ticker"]] = []
    blocks = []
    for mkey, label, higher_better in _HEADLINE_CHARTS:
        m_def = METRICS_BY_KEY.get(mkey)
        if not m_def:
            continue
        fld = m_def.get("fdic_field")   # direct FDIC field → sparkline available
        pts = [(p["ticker"], p.get(mkey)) for p in display_peers
               if isinstance(p.get(mkey), (int, float))]
        if len(pts) < 2:
            continue
        fmt = m_def.get("format", "number")
        dec = m_def.get("decimals", 2)
        vals = [v for _, v in pts]
        median = pd.Series(vals).median()
        max_abs = max(abs(v) for v in vals) or 1.0
        best_tk = (max if higher_better else min)(pts, key=lambda x: x[1])[0]
        pts_sorted = sorted(pts, key=lambda x: x[1], reverse=higher_better)
        bars = []
        for tk, v in pts_sorted:
            w = max(4, round(100 * abs(v) / max_abs))
            color = "var(--brand-primary)" if tk == best_tk else "#a9bbdc"
            vs = _html.escape(format_value(v, fmt, dec))
            spark = ""
            if fld:
                ser = list(reversed([r.get(fld) for r in (hist.get(tk) or [])[:8]]))
                spark = _sparkline_svg(ser)
            bars.append(
                f'<div style="display:flex;align-items:center;gap:5px;margin:2px 0;">'
                f'<span style="width:38px;font-size:0.7rem;color:var(--text-secondary);'
                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'
                f'{_html.escape(tk)}</span>'
                f'<div style="flex:1;min-width:24px;background:var(--grid-head-bg);'
                f'border-radius:0;"><div style="height:11px;width:{w}%;'
                f'background:{color};border-radius:0;"></div></div>'
                f'<span style="width:44px;text-align:right;font-size:0.7rem;">{vs}</span>'
                f'{spark}</div>')
        note = "" if higher_better else " · lower better"
        med_s = _html.escape(format_value(median, fmt, dec))
        blocks.append(
            f'<div style="margin-bottom:13px;">'
            f'<div style="font-size:0.78rem;margin-bottom:3px;">{_html.escape(label)} '
            f'<span style="color:var(--text-tertiary);">· med {med_s}{note}</span></div>'
            f'{"".join(bars)}</div>')
    if not blocks:
        return
    st.markdown(
        '<div style="font-size:0.68rem;letter-spacing:.04em;text-transform:uppercase;'
        'color:var(--text-tertiary);margin:2px 0 8px;">At a glance '
        '<span style="text-transform:none;letter-spacing:0;">'
        '· bar = current · line = 8q trend</span></div>'
        + "".join(blocks),
        unsafe_allow_html=True,
    )


def _compare_export_bytes(cohort: list[dict], categories: list[str]):
    """Build (xlsx, csv) of the WHOLE cohort as a SORTABLE sheet: banks in rows,
    each metric (across the selected categories) a RAW-number column + an overall-
    score column. The xlsx has AutoFilter, a frozen header + ticker/bank columns,
    and per-column number formats so Excel sorts numerically. No colors."""
    import io
    import csv as _csv

    mcols = []   # (key, label, format) — metrics with any data in the cohort
    for cat in categories:
        for mkey in CATEGORY_METRICS.get(cat, []):
            m_def = METRICS_BY_KEY.get(mkey)
            if m_def and any(isinstance(b.get(mkey), (int, float)) for b in cohort):
                mcols.append((mkey, m_def.get("label", mkey),
                              m_def.get("format", "number")))
    scores = _peer_scores(cohort, cohort, categories)
    headers = ["Ticker", "Bank", "Overall score"] + [lbl for _, lbl, _ in mcols]
    rows = []
    for b in cohort:
        tk = b.get("ticker")
        sc = scores.get(tk, {}).get("score")
        r = [tk, get_name(tk), (round(sc) if sc is not None else None)]
        for mkey, _, _ in mcols:
            v = b.get(mkey)
            r.append(v if isinstance(v, (int, float)) else None)
        rows.append(r)

    sbuf = io.StringIO()
    _w = _csv.writer(sbuf)
    _w.writerow(headers)
    _w.writerows(rows)
    csv_bytes = sbuf.getvalue().encode("utf-8")

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Peer comparison"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for rr in range(2, ws.max_row + 1):
        ws.cell(rr, 3).number_format = "0"   # score (integer)
    for ci, (_, _, fmt) in enumerate(mcols, start=4):
        nf = {"pct": '0.00"%"', "ratio": '0.00"x"'}.get(fmt, '#,##0.00')
        for rr in range(2, ws.max_row + 1):
            ws.cell(rr, ci).number_format = nf
    ws.freeze_panes = "C2"   # freeze header row + Ticker/Bank
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 13
    for ci in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(
            22, max(10, len(headers[ci - 1]) + 2))
    xbuf = io.BytesIO()
    wb.save(xbuf)
    return xbuf.getvalue(), csv_bytes


def _render_metrics_table(cohort: list[dict], display_peers: list[dict],
                          categories: list[str], scores: dict | None = None,
                          delta: bool = False):
    """Dense side-by-side table — banks (the display subset) in columns, metrics in
    rows grouped under category section headers, EVERY metric per selected category.
    Percentile color and Peer Median resolve against the FULL ``cohort``. When
    ``delta`` is set, bank cells show the signed difference from the peer median.
    ``scores`` (per-ticker composite) pins an Overall-score row at the top."""
    scores = scores or {}
    tickers = [m["ticker"] for m in display_peers]   # columns
    style_map = {}                                   # (metric_key, ticker) → color
    sections = []                                    # [(category, [row, …]), …]

    for cat in categories:
        cat_rows = []
        cat_effs = {t: [] for t in tickers}   # effective percentiles for the subtotal
        for mkey in CATEGORY_METRICS.get(cat, []):
            m_def = METRICS_BY_KEY.get(mkey)
            if not m_def:
                continue
            # Percentile basis = the full cohort, not just the displayed banks.
            numeric = [b.get(mkey) for b in cohort
                       if isinstance(b.get(mkey), (int, float))]
            if not numeric:
                continue
            peer_median = pd.Series(numeric).median()
            higher_better = m_def.get("color_rule") == "higher_better"
            lower_better = m_def.get("color_rule") == "lower_better"
            fmt = m_def.get("format", "number")
            dec = m_def.get("decimals", 2)
            row = {"Metric": m_def["label"], "_mkey": mkey, "_cat": cat}
            for d in display_peers:
                t = d["ticker"]
                v = d.get(mkey)
                if v is None or not isinstance(v, (int, float)):
                    row[t] = "—"
                elif delta:
                    dv = v - peer_median
                    row[t] = ("+" if dv >= 0 else "−") + format_value(
                        abs(dv), fmt, dec)
                else:
                    row[t] = format_value(v, fmt, dec)
                pct = compute_peer_percentile(v, numeric)   # vs full cohort
                if higher_better:
                    style_map[(mkey, t)] = _percentile_color(pct, True)
                elif lower_better:
                    style_map[(mkey, t)] = _percentile_color(pct, False)
                else:
                    style_map[(mkey, t)] = ""
                if (higher_better or lower_better) and pct is not None \
                        and isinstance(v, (int, float)):
                    cat_effs[t].append(pct if higher_better else (100 - pct))
            row["Peer Median"] = format_value(peer_median, fmt, dec)
            cat_rows.append(row)
        if cat_rows:
            # Per-category subtotal = each bank's mean EFFECTIVE percentile within
            # the category (the only unit-neutral "average"); {ticker: (str, avg)}.
            sub_vals = {}
            for t in tickers:
                effs = cat_effs[t]
                avg = (sum(effs) / len(effs)) if effs else None
                sub_vals[t] = (str(round(avg)) if avg is not None else "—", avg)
            sections.append((cat, cat_rows, sub_vals))

    if not sections:
        st.warning("No metrics to display for the selected categories.")
        return

    n_cols = len(tickers) + 2
    head = ('<th class="nm">Metric</th>'
            + "".join(
                f'<th><a class="lnk tk" href="?bank={_html.escape(t, quote=True)}" '
                f'target="_self">{_html.escape(t)}</a></th>' for t in tickers)
            + '<th>Peer Median</th>')

    # Pinned "Overall score" row (composite percentile) at the top of the body.
    ov_cells = ['<td class="nm">Overall score</td>']
    for t in tickers:
        _sc = scores.get(t, {}).get("score")
        _sty = _percentile_color(_sc, True) if _sc is not None else ""
        _style = f' style="{_sty}"' if _sty else ""
        ov_cells.append(
            f'<td class="num"{_style}>{round(_sc) if _sc is not None else "—"}</td>')
    ov_cells.append('<td class="num med">50</td>')
    body_rows = ['<tr class="overall">' + "".join(ov_cells) + "</tr>"]

    for cat, cat_rows, sub_vals in sections:
        body_rows.append(
            f'<tr class="sec"><td class="nm" colspan="{n_cols}">'
            f'{_html.escape(cat)}</td></tr>')
        for row in cat_rows:
            mkey = row["_mkey"]
            cells = [f'<td class="nm">{_html.escape(str(row["Metric"]))}</td>']
            for t in tickers:
                sty = style_map.get((mkey, t), "")
                style = f' style="{sty}"' if sty else ""
                cells.append(
                    f'<td class="num"{style}>{_html.escape(str(row.get(t, "—")))}</td>')
            cells.append(
                f'<td class="num med">{_html.escape(str(row.get("Peer Median", "—")))}</td>')
            body_rows.append("<tr>" + "".join(cells) + "</tr>")
        # Category subtotal: mean effective percentile per bank (colored higher=better).
        scells = ['<td class="nm">Avg percentile</td>']
        for t in tickers:
            disp, avg = sub_vals[t]
            sty = _percentile_color(avg, True) if avg is not None else ""
            style = f' style="{sty}"' if sty else ""
            scells.append(f'<td class="num"{style}>{_html.escape(disp)}</td>')
        scells.append('<td class="num med">50</td>')
        body_rows.append('<tr class="subtot">' + "".join(scells) + "</tr>")

    # Table on the LEFT; a rail of headline bar charts on the RIGHT fills the
    # space (the table scrolls within its own column when there are many banks).
    tbl_col, chart_col = st.columns([3, 1.4])
    with tbl_col:
        st.markdown(
            "<style>"
            ".cmp-wrap{max-height:640px;overflow:auto;border:0.5px solid var(--grid-head);}"
            ".cmp-wrap thead th{position:sticky;top:0;z-index:3;}"
            ".cmp-wrap td.nm,.cmp-wrap th.nm{text-align:left;}"
            ".cmp-wrap td.med{color:var(--text-secondary);font-weight:600;}"
            ".cmp-wrap tr.sec td{position:sticky;left:0;background:var(--grid-head-bg);"
            "color:var(--brand-primary);text-transform:uppercase;letter-spacing:.04em;"
            "font-size:0.68rem;font-weight:600;padding:5px 10px;}"
            ".cmp-wrap tr.subtot td{border-top:0.5px solid var(--grid-head);"
            "border-bottom:0.5px solid var(--grid-head);font-weight:600;}"
            ".cmp-wrap tr.subtot td.nm{font-style:italic;color:var(--text-secondary);}"
            ".cmp-wrap tr.overall td{font-weight:700;"
            "border-bottom:1px solid var(--grid-head);}"
            ".cmp-wrap tr.overall td.nm{color:var(--brand-primary);}"
            ".cmp-wrap thead th a.tk{color:inherit;text-decoration:none;font-weight:600;}"
            ".cmp-wrap thead th a.tk:hover{color:var(--brand-primary);"
            "text-decoration:underline;}"
            "</style>"
            f'<div class="cmp-wrap"><table class="ksk-grid">'
            f'<thead><tr>{head}</tr></thead><tbody>{"".join(body_rows)}</tbody></table></div>',
            unsafe_allow_html=True,
        )

        # Export: the WHOLE cohort as a sortable sheet (banks in rows, raw numbers).
        # Built on dialog-open so the per-bank scoring runs only on demand.
        @st.dialog("Export comparison")
        def _cmp_export_dialog():
            st.caption(f"All {len(cohort)} banks in scope × {len(categories)} "
                       "categories — banks in rows, every metric a sortable column "
                       "(raw numbers; the .xlsx has AutoFilter + frozen header).")
            with st.spinner("Building…"):
                _xlsx, _csv_bytes = _compare_export_bytes(cohort, categories)
            e1, e2 = st.columns(2)
            with e1:
                st.download_button(
                    "Excel (sortable)", _xlsx, file_name="peer_comparison.xlsx",
                    mime="application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet",
                    use_container_width=True, key="cmp_xlsx_dl")
            with e2:
                st.download_button(
                    "CSV", _csv_bytes, file_name="peer_comparison.csv",
                    mime="text/csv", use_container_width=True, key="cmp_csv_dl")
        if st.button("Export", key="cmp_export_btn"):
            _cmp_export_dialog()

        # ── Legend — same scale that colors the cells ──────────────────
        chips = "".join(
            f'<div style="background:{bg}; padding:4px 10px; border-radius:0; '
            f'color:{fg};">{"<b>" + label + "</b>" if bold else label}</div>'
            for _floor, label, bg, fg, bold in _PCT_SCALE
        )
        st.markdown(
            '<div style="display:flex; gap:12px; margin-top:8px; flex-wrap:wrap; '
            f'font-size:0.8rem; align-items:center;">'
            f'<span style="color:var(--text-secondary);">Percentile vs full scope:</span>'
            f'{chips}</div>',
            unsafe_allow_html=True,
        )
    with chart_col:
        _render_headline_charts(display_peers)


# ── Scatter Plots ────────────────────────────────────────────────────

_CURATED_SCATTERS = [
    {"name": "Profitability vs Efficiency",
     "x": "efficiency_ratio", "y": "roatce",
     "x_label": "Efficiency Ratio (%)", "y_label": "ROATCE (%)",
     "x_invert": True,  # lower efficiency = better
     "quadrants": {"TR": "Best-in-class", "TL": "High ROE but inefficient",
                   "BR": "Efficient but low ROE", "BL": "Laggards"}},
    {"name": "Margin vs Capital",
     "x": "cet1_ratio", "y": "nim",
     "x_label": "CET1 Ratio (%)", "y_label": "NIM (%)",
     "quadrants": {"TR": "Strong margin + capital", "TL": "Strong margin, thin capital",
                   "BR": "Over-capitalized, weak margin", "BL": "Weak across both"}},
    {"name": "Valuation vs Profitability",
     "x": "roatce", "y": "ptbv_ratio",
     "x_label": "ROATCE (%)", "y_label": "P/TBV",
     "quadrants": {"TR": "Richly valued, high-ROE", "TL": "Rich multiple, low-ROE",
                   "BR": "Cheap, high-ROE (value)", "BL": "Cheap, low-ROE (value trap)"}},
    {"name": "Growth vs Credit",
     "x": "npl_ratio", "y": "dep_qoq_growth",
     "x_label": "NPL Ratio (%)", "y_label": "Deposit QoQ Growth (%)",
     "x_invert": True,
     "quadrants": {"TR": "Clean growth", "TL": "Deteriorating + growing",
                   "BR": "Shrinking but clean", "BL": "Shrinking + deteriorating"}},
]


def _render_peer_scatters(selected_peers: list[dict]):
    """Render curated preset scatters + custom 2-axis picker."""
    if len(selected_peers) < 3:
        st.info("Need at least 3 banks in the peer group for meaningful scatter plots.")
        return

    # ── Curated presets (2x2 grid) ────────────────────────────────────
    st.markdown("##### Curated Presets")
    col_l, col_r = st.columns(2)
    for i, preset in enumerate(_CURATED_SCATTERS):
        container = col_l if i % 2 == 0 else col_r
        with container:
            fig = _build_scatter(selected_peers, preset)
            if fig is not None:
                st.plotly_chart(fig, use_container_width=True)

    # ── Custom picker ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("##### Custom Scatter")

    # Build list of pickable metrics — numeric ones only
    numeric_metrics = [
        (k, METRICS_BY_KEY[k]["label"])
        for k in METRICS_BY_KEY
        if METRICS_BY_KEY[k].get("format") in ("pct", "ratio", "currency", "number", "millions", "billions", "dollars_auto")
    ]
    numeric_metrics.sort(key=lambda x: x[1])
    metric_keys = [k for k, _ in numeric_metrics]
    metric_labels = {k: lbl for k, lbl in numeric_metrics}

    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        default_x = "efficiency_ratio" if "efficiency_ratio" in metric_keys else metric_keys[0]
        x_key = st.selectbox(
            "X axis", metric_keys,
            index=metric_keys.index(default_x) if default_x in metric_keys else 0,
            format_func=lambda k: metric_labels.get(k, k),
            key="custom_scatter_x",
        )
    with cc2:
        default_y = "roatce" if "roatce" in metric_keys else metric_keys[1]
        y_key = st.selectbox(
            "Y axis", metric_keys,
            index=metric_keys.index(default_y) if default_y in metric_keys else 1,
            format_func=lambda k: metric_labels.get(k, k),
            key="custom_scatter_y",
        )
    with cc3:
        size_metric = st.selectbox(
            "Bubble size", ["(uniform)"] + metric_keys,
            index=(metric_keys.index("total_assets") + 1) if "total_assets" in metric_keys else 0,
            format_func=lambda k: metric_labels.get(k, "Uniform") if k != "(uniform)" else "Uniform",
            key="custom_scatter_size",
        )

    custom_preset = {
        "name": f"{metric_labels.get(y_key, y_key)} vs {metric_labels.get(x_key, x_key)}",
        "x": x_key, "y": y_key,
        "x_label": metric_labels.get(x_key, x_key),
        "y_label": metric_labels.get(y_key, y_key),
        "size": size_metric if size_metric != "(uniform)" else None,
    }

    fig = _build_scatter(selected_peers, custom_preset, height=420)
    if fig is not None:
        st.plotly_chart(fig, use_container_width=True)


def _build_scatter(peers: list[dict], preset: dict, height: int = 320):
    """Build a single scatter plot from a preset config."""
    import plotly.graph_objects as go
    from utils.chart_style import (
        apply_standard_layout, COLOR_PRIMARY, COLOR_NEUTRAL, COLOR_GREY_LIGHT,
    )

    x_key = preset["x"]
    y_key = preset["y"]
    size_key = preset.get("size")

    # Collect data
    points = []
    for p in peers:
        x = p.get(x_key)
        y = p.get(y_key)
        if x is None or y is None:
            continue
        size = p.get(size_key) if size_key else None
        points.append({
            "ticker": p["ticker"],
            "name": get_name(p["ticker"]),
            "x": x, "y": y,
            "size": size,
        })

    if len(points) < 2:
        return None

    x_vals = [pt["x"] for pt in points]
    y_vals = [pt["y"] for pt in points]

    # Size: scale to [12, 40] bubble area
    if size_key and any(pt.get("size") is not None for pt in points):
        raw_sizes = [pt.get("size") or 0 for pt in points]
        max_s = max(abs(s) for s in raw_sizes) or 1
        sizes = [12 + 28 * (abs(s) / max_s) for s in raw_sizes]
    else:
        sizes = [18] * len(points)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=x_vals, y=y_vals,
        mode="markers+text",
        text=[pt["ticker"] for pt in points],
        textposition="top center",
        textfont=dict(size=10, color=COLOR_NEUTRAL),
        marker=dict(
            size=sizes,
            color=COLOR_PRIMARY,
            opacity=0.65,
            line=dict(color=COLOR_PRIMARY, width=1),
        ),
        customdata=[[pt["name"], pt["x"], pt["y"]] for pt in points],
        hovertemplate=(
            "<b>%{text}</b> — %{customdata[0]}<br>"
            f"{preset['x_label']}: %{{x:.2f}}<br>"
            f"{preset['y_label']}: %{{y:.2f}}<extra></extra>"
        ),
    ))

    # Median crosshairs
    x_med = pd.Series(x_vals).median()
    y_med = pd.Series(y_vals).median()
    fig.add_vline(x=x_med, line_color=COLOR_GREY_LIGHT, line_width=1, line_dash="dot")
    fig.add_hline(y=y_med, line_color=COLOR_GREY_LIGHT, line_width=1, line_dash="dot")

    # Invert x-axis if lower-is-better
    apply_standard_layout(
        fig, title=preset["name"],
        height=height,
        xaxis_title=preset["x_label"], yaxis_title=preset["y_label"],
        show_legend=False, hovermode="closest",
    )
    if preset.get("x_invert"):
        fig.update_xaxes(autorange="reversed")

    return fig


# ── Radar Chart ──────────────────────────────────────────────────────

_RADAR_METRICS = [
    # Normalized ROATCE (one-time spikes winsorized) for a fair peer comparison
    # — the single-quarter "roatce" can spike (e.g. CARE's loan-recovery quarter
    # → ~71%) and distort the radar.
    ("roatce_normalized", "ROATCE", True),   # higher better
    ("nim", "NIM", True),
    ("efficiency_ratio", "Efficiency", False),  # lower better — will invert
    ("cet1_ratio", "CET1", True),
    ("npl_ratio", "NPL", False),              # lower better
    ("nonint_dep_pct", "Non-Int Dep %", True),
    ("tbv_cagr_1y", "TBV CAGR", True),
    ("ptbv_ratio", "P/TBV", False),           # lower better (cheaper)
]


def _render_peer_radar(selected_peers: list[dict]):
    """Render a radar chart comparing banks on 8 key metrics by percentile rank."""
    import plotly.graph_objects as go
    from utils.chart_style import (
        apply_standard_layout, CATEGORICAL_PALETTE, CHART_HEIGHT_FULL,
        COLOR_NEUTRAL, _GRID_COLOR, _AXIS_COLOR, _BG_SURFACE,
    )

    if len(selected_peers) < 2:
        st.info("Need at least 2 banks for radar comparison.")
        return

    # Bank picker — up to 5 banks overlay
    tickers = [p["ticker"] for p in selected_peers]
    default = tickers[:min(4, len(tickers))]
    picked = st.multiselect(
        "Banks to overlay (up to 5)",
        tickers,
        default=default,
        max_selections=5,
        key="radar_picker",
    )

    if not picked:
        st.info("Pick at least 1 bank above to render the radar.")
        return

    # Compute percentile for each metric across the entire peer group
    # (always use full peer group for percentile, even if user picked subset)
    metric_data = []
    for mkey, label, higher_better in _RADAR_METRICS:
        values = [p.get(mkey) for p in selected_peers]
        numeric = [v for v in values if isinstance(v, (int, float)) and v is not None]
        if not numeric:
            continue
        metric_data.append({
            "key": mkey, "label": label, "higher_better": higher_better,
            "values": values, "numeric": numeric,
        })

    if len(metric_data) < 3:
        st.info("Not enough metrics with data for a meaningful radar.")
        return

    categories = [m["label"] for m in metric_data]

    fig = go.Figure()
    # Series colors come from the shared categorical palette; the translucent
    # fill is derived from each palette hex (no ad-hoc series hexes).
    colors = CATEGORICAL_PALETTE
    def _fill(hex_color: str, alpha: float = 0.14) -> str:
        h = hex_color.lstrip("#")
        r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
        return f"rgba({r}, {g}, {b}, {alpha})"
    fill_colors = [_fill(c) for c in colors]

    for i, ticker in enumerate(picked):
        bank = next((p for p in selected_peers if p["ticker"] == ticker), None)
        if not bank:
            continue

        # Compute percentile rank for each metric (0-100)
        r_values = []
        for m in metric_data:
            val = bank.get(m["key"])
            if val is None:
                r_values.append(0)
                continue
            numeric = m["numeric"]
            below = sum(1 for v in numeric if v < val)
            pct = (below / len(numeric)) * 100 if numeric else 0
            # Invert if lower-is-better
            effective_pct = pct if m["higher_better"] else (100 - pct)
            r_values.append(effective_pct)

        # Close the polygon
        r_closed = r_values + [r_values[0]]
        cats_closed = categories + [categories[0]]

        color = colors[i % len(colors)]
        fill = fill_colors[i % len(fill_colors)]
        fig.add_trace(go.Scatterpolar(
            r=r_closed, theta=cats_closed,
            fill="toself",
            fillcolor=fill,
            name=ticker,
            line=dict(color=color, width=2),
            opacity=0.9,
        ))

    # Standard chrome (bg / font / legend-below / tight margins / title);
    # the polar block (which apply_standard_layout doesn't model) is layered
    # on after, using the shared chart tokens.
    apply_standard_layout(
        fig, title="Peer Radar — Percentile Rank (0–100)",
        height=CHART_HEIGHT_FULL, show_legend=True, hovermode="closest",
    )
    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True, range=[0, 100],
                tickfont=dict(size=10, color=COLOR_NEUTRAL),
                gridcolor=_GRID_COLOR,
                linecolor=_AXIS_COLOR,
            ),
            angularaxis=dict(
                tickfont=dict(size=11, color=COLOR_NEUTRAL),
                gridcolor=_GRID_COLOR,
                linecolor=_AXIS_COLOR,
            ),
            bgcolor=_BG_SURFACE,
        ),
    )
    st.plotly_chart(fig, use_container_width=True)

    st.caption(
        "Each axis = percentile rank of that metric within the full peer group "
        "(higher = better for this metric; lower-is-better metrics inverted). "
        "Bigger polygon = stronger overall."
    )
